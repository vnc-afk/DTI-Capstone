import datetime
from itertools import chain
from urllib.parse import quote
from django.urls import reverse
from django.views import View
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from django.http import HttpResponse, HttpResponseRedirect
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from ..mappings import EXPORT_MODEL_MAP, UPLOAD_MODEL_MAP
from ..utils.excel_view_helpers import get_model_from_sheet, get_user_from_full_name, normalize_header, normalize_sheet_name, shorten_name, to_title
from django.contrib.auth.mixins import LoginRequiredMixin
from ..mixins.permissions_mixins import UserRoleMixin
from django.shortcuts import render
from django.contrib import messages
from django.db import IntegrityError, transaction
import pandas as pd
from openpyxl import load_workbook
import os
from ..models import (
    ChecklistEvaluationSheet,
    InspectionValidationReport,
    OrderOfPayment,
    PersonalDataSheet,
    SalesPromotionPermitApplication,
    ServiceRepairAccreditationApplication,
)


class ExportDocumentsExcelView(LoginRequiredMixin, View):
    """
    Export selected documents (via checkboxes) or all filtered documents to Excel.
    Preserves grouped formatting, headers, styling, and filename date ranges.
    """

    def post(self, request, *args, **kwargs):
        user = request.user
        selected_docs = request.POST.getlist("documents")

        # --- Collect documents by model ---
        grouped = {}
        if selected_docs:
            # Use checkbox selection
            grouped_ids = {}
            for doc in selected_docs:
                try:
                    model_name, pk = doc.split(":")
                    grouped_ids.setdefault(model_name.lower(), []).append(pk)
                except ValueError:
                    continue

            for model_name, ids in grouped_ids.items():
                model = EXPORT_MODEL_MAP.get(model_name)
                if not model:
                    continue
                qs = UserRoleMixin.get_queryset_or_all(model, user).filter(pk__in=ids)
                if qs.exists():
                    grouped[model._meta.verbose_name_plural.title()] = list(qs)
        else:
            # If no checkboxes, export all (apply UserRoleMixin)
            for model in EXPORT_MODEL_MAP.values():
                qs = UserRoleMixin.get_queryset_or_all(model, user)
                if qs.exists():
                    grouped[model._meta.verbose_name_plural.title()] = list(qs)

        # --- Create workbook ---
        wb = Workbook()
        wb.remove(wb.active)

        for model_name, objs in grouped.items():
            if not objs:
                continue
            # use your helper instead of hard truncation
            sheet_title = shorten_name(model_name)
            ws = wb.create_sheet(title=sheet_title)

            fields = objs[0]._meta.fields
            headers = [f.verbose_name.title() for f in fields]
            ws.append(headers)

            # Style headers
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

            # Rows
            for obj in objs:
                row = []
                for f in fields:
                    val = getattr(obj, f.name, "")
                    if isinstance(val, datetime.datetime):
                        val = val.date()
                    if val is None:
                        val = ""
                    row.append(to_title(str(val)))
                ws.append(row)

            # Auto-size columns
            for col_num, col_cells in enumerate(ws.columns, 1):
                max_length = 0
                column = get_column_letter(col_num)
                for cell in col_cells:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[column].width = min(max_length + 2, 50)

            # Freeze header
            ws.freeze_panes = "A2"

        # --- Filename with date range ---
        dates = []
        for objs in grouped.values():
            for doc in objs:
                d = getattr(doc, "date_filed", None) or getattr(doc, "date", None)
                if d:
                    if isinstance(d, datetime.datetime):
                        d = d.date()
                    dates.append(d)
        min_date, max_date = (min(dates), max(dates)) if dates else (None, None)
        date_part = f"{min_date}_{max_date}" if min_date and max_date else ""

        filename = f"Documents_{date_part}.xlsx" if date_part else "Documents.xlsx"

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        safe_filename = quote(filename)
        response["Content-Disposition"] = f"attachment; filename={safe_filename}; filename*=UTF-8''{safe_filename}"
        wb.save(response)
        return response

class UploadExcelView(View):
    def post(self, request, *args, **kwargs):
        files = request.FILES.getlist("files")
        
        if not files:
            messages.error(request, "No files were uploaded. Please select an Excel file to continue.")
            return HttpResponseRedirect(reverse("documents:all-documents"))

        file_reports = []

        for file in files:
            report = {
                "filename": file.name,
                "filesize": round(file.size / (1024 * 1024), 2),  # MB
                "created_at": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "total_records": 0,
                "imported_records": 0,
                "failed_records": 0,
                "documents": [],
            }

            ext = os.path.splitext(file.name)[1].lower()
            # Skip temporary Excel files (those starting with ~$)
            if file.name.startswith("~$"):
                messages.error(request, f"Temporary file {file.name} is not a valid Excel file.")
                continue

            if ext not in [".xlsx", ".xls"]:
                messages.error(request, f"{file.name} was not uploaded. Only Excel files are allowed.")
                continue

            try:
                wb = load_workbook(file)
            except Exception:
                messages.error(request, f"Could not open {file.name}.")
                continue

            for sheetname in wb.sheetnames:
                model = get_model_from_sheet(sheetname)
                if not model:
                    continue

                ws = wb[sheetname]
                headers = [normalize_header(str(c.value).strip()) if c.value else "" for c in ws[1]]

                field_map = {}
                for f in model._meta.fields:
                    opts = {normalize_header(f.verbose_name), normalize_header(f.name)}
                    for h in headers:
                        if h in opts:
                            field_map[h] = f.name

                row_number = 1  # This will track the actual Excel row number
                for row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    if not any(row):
                        continue

                    # Initialize record with display info even if it fails
                    record = {
                        "display": "",  # Will be set later based on success/failure
                        "model": model._meta.verbose_name,
                        "status": "Imported",
                        "invalid_fields": [],
                        "user": None,
                        "row_number": row_index - 1,  # Use enumerate index for exact Excel row
                    }

                    data = {}
                    for header, value in zip(headers, row):
                        field_name = field_map.get(header)
                        if not field_name:
                            continue

                        if isinstance(value, datetime.datetime):
                            value = value.date()
                        if value in ("", None):
                            value = None

                        if field_name == "user" and value:
                            user_obj = get_user_from_full_name(str(value))
                            if not user_obj:
                                record["status"] = "Failed"
                                record["invalid_fields"].append("invalid user")
                                record["display"] = f"Row {record['row_number']}"  # Set display immediately
                                continue
                            data[field_name] = user_obj
                            record["user"] = str(user_obj)
                        else:
                            data[field_name] = value

                    try:
                        if data and record["status"] == "Imported":
                            with transaction.atomic():
                                obj = model.objects.create(**data)
                                record["display"] = str(obj)
                                record["date_created"] = getattr(obj, "created_at", datetime.datetime.now()).strftime("%Y-%m-%d %H:%M:%S")
                            report["imported_records"] += 1
                        else:
                            if record["status"] == "Failed":
                                # For failed records, show row number
                                record["display"] = f"Row {record['row_number']}"
                                report["failed_records"] += 1
                    except Exception as e:
                        record["status"] = "Failed"
                        # For failed records, show row number
                        record["display"] = f"Row {record['row_number']}"
                        
                        error_msg = str(e).lower()
                        
                        # Check if it's a NOT NULL constraint error
                        if 'not null' in error_msg or 'null constraint' in error_msg:
                            # Extract field name from error message
                            # Common patterns: "NOT NULL constraint failed: table.field"
                            # or "null value in column "field" violates not-null constraint"
                            import re
                            
                            # Try to extract field name from different error message formats
                            field_match = None
                            
                            # Pattern 1: "NOT NULL constraint failed: table.field"
                            pattern1 = r'NOT NULL constraint failed: \w+\.(\w+)'
                            field_match = re.search(pattern1, error_msg, re.IGNORECASE)
                            
                            if not field_match:
                                # Pattern 2: 'null value in column "field" violates'
                                pattern2 = r'null value in column ["\'](\w+)["\']'
                                field_match = re.search(pattern2, error_msg, re.IGNORECASE)
                            
                            if not field_match:
                                # Pattern 3: "(1048, "Column 'field' cannot be null")"
                                pattern3 = r"Column ['\"](\w+)['\"] cannot be null"
                                field_match = re.search(pattern3, error_msg, re.IGNORECASE)
                            
                            if field_match:
                                field_name = field_match.group(1)
                                # Get the verbose name if available
                                try:
                                    field_obj = model._meta.get_field(field_name)
                                    field_display = field_obj.verbose_name or field_name
                                except:
                                    field_display = field_name
                                
                                record["invalid_fields"].append(f"missing field ({field_display})")
                            else:
                                record["invalid_fields"].append(f"missing required field")
                        else:
                            # For other types of errors, show the original message
                            record["invalid_fields"].append(str(e))
                        
                        report["failed_records"] += 1

                    report["total_records"] += 1
                    
                    # Format invalid fields for display
                    if record["invalid_fields"]:
                        if len(record["invalid_fields"]) <= 2:
                            record["invalid_fields_display"] = ", ".join(record["invalid_fields"])
                        else:
                            first_two = ", ".join(record["invalid_fields"][:2])
                            remaining_count = len(record["invalid_fields"]) - 2
                            record["invalid_fields_display"] = f"{first_two} + {remaining_count} other invalid/missing fields"
                    else:
                        record["invalid_fields_display"] = "None"
                    
                    report["documents"].append(record)

            file_reports.append(report)

        # Store in session safely (JSON-serializable)
        request.session["file_reports"] = file_reports

        return HttpResponseRedirect(reverse("upload-report"))

class UploadReportView(View):
    template_name = "documents/excel_templates/upload_report.html"

    def get(self, request, *args, **kwargs):
        # Pull file_reports from session, default to an empty list
        file_reports = request.session.get("file_reports", [])

        context = {
            "file_reports": file_reports,
        }

        return render(request, self.template_name, context)

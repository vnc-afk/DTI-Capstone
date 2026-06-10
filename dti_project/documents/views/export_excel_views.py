from django.http import HttpResponse
from django.views import View
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from documents.models import CollectionReportItem
from datetime import datetime
from decimal import Decimal

class ExportCollectionReportView(View):
    """Export collection report to Excel with DTI format"""

    def get(self, request, *args, **kwargs):
        # Get the same items that appear in the UI
        collections = self.get_page_collections(request)

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Collections Report"

        # Format worksheet
        self.format_report(ws, collections)

        # Prepare response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = (
            f'attachment; filename=collection_report_'
            f'{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )

        wb.save(response)
        return response

    def get_page_collections(self, request):
        """
        Get the same items shown in your Collection Report page.
        Update filtering logic as needed.
        """

        collections = CollectionReportItem.objects.all().order_by('date', 'number')
        return collections

    def format_report(self, ws, collections):
        """Format Excel to match DTI report layout"""

        # Define styles
        title_font = Font(name='Arial', size=11, bold=True)
        header_font = Font(name='Arial', size=9, bold=True)
        normal_font = Font(name='Arial', size=9)

        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align = Alignment(horizontal='left', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

        # Set column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 35
        ws.column_dimensions['E'].width = 40
        ws.column_dimensions['F'].width = 15

        # Title
        ws.merge_cells('A1:F1')
        ws['A1'] = 'REPORT OF COLLECTIONS AND DEPOSITS'
        ws['A1'].font = title_font
        ws['A1'].alignment = center_align
        ws['A1'].border = thin_border

        # Office Name
        ws.merge_cells('A3:F3')
        ws['A3'] = 'DTI Albay Provincial Office'
        ws['A3'].font = Font(name='Arial', size=10, bold=True)
        ws['A3'].alignment = center_align

        # Date / Report No.
        # === Get parent report header ===
        report = None
        if collections:
            report = collections.first().collection_reports.first()
            collections = report.report_items.all()  # Reset to report's items for consistency

        report_date = (
            report.report_collection_date.strftime('%d-%b-%y')
            if report and report.report_collection_date else ""
        )       

        report_no = report.report_no if report and report.report_no else ""

            # === Write report date ===
        ws.merge_cells('A4:E4')
        ws['A4'] = report_date
        ws['A4'].font = normal_font
        ws['A4'].alignment = center_align

            # === Write report number ===
        ws['F4'] = f"Report No. {report_no}"
        ws['F4'].font = normal_font
        ws['F4'].alignment = right_align


        # Column Headers
        headers = ['Date', 'Official Receipt\nNumber', 'Responsibility\nCenter Code',
                   'Payor', 'Particulars', 'Amount']

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col)
            cell.value = header
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
            cell.fill = header_fill

        # Data Rows
        current_row = 7
        total_amount = Decimal('0.00')

        for item in collections:
            ws.cell(row=current_row, column=1, value=item.date.strftime('%d-%b-%y'))
            ws.cell(row=current_row, column=2, value=item.number)
            ws.cell(row=current_row, column=3, value=item.rc_code or '')
            ws.cell(row=current_row, column=4, value=item.payor)
            ws.cell(row=current_row, column=5, value=item.particulars)
            ws.cell(row=current_row, column=6, value=float(item.amount or 0))

            # Apply formatting
            for col in range(1, 7):
                cell = ws.cell(row=current_row, column=col)
                cell.font = normal_font
                cell.border = thin_border

                if col == 6:
                    cell.alignment = right_align
                    cell.number_format = '#,##0.00'
                elif col == 1:
                    cell.alignment = center_align
                else:
                    cell.alignment = left_align

            total_amount += Decimal(str(item.amount or 0))
            current_row += 1

        # (the rest of your summary + certification formatting remains unchanged)
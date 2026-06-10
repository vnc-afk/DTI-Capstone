import datetime
from io import BytesIO
from itertools import chain
import json
import re
import threading
import time
from urllib.parse import quote
import uuid
from django.urls import reverse
from django.views import View
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from django.http import HttpResponse, HttpResponseRedirect, StreamingHttpResponse
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from ..models.collection_models import CollectionReport, CollectionReportItem
from ..mappings import EXPORT_MODEL_MAP, UPLOAD_MODEL_MAP
from ..utils.excel_view_helpers import get_model_from_sheet, get_user_from_full_name, normalize_header, normalize_sheet_name, shorten_name, to_title
from django.contrib.auth.mixins import LoginRequiredMixin
from ..mixins.permissions_mixins import UserRoleMixin
from django.shortcuts import redirect, render
from django.contrib import messages
from django.db import IntegrityError, transaction
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel
from django.utils.dateparse import parse_date
from django.http import JsonResponse
from django.core.cache import cache
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

import os
from ..models import (
    ChecklistEvaluationSheet,
    InspectionValidationReport,
    OrderOfPayment,
    PersonalDataSheet,
    SalesPromotionPermitApplication,
    ServiceRepairAccreditationApplication,
)
    
class UploadExcelView(View):
    """
    Initial upload endpoint - stores files in cache and returns session ID
    """
    
    @staticmethod
    def normalize_header(value):
        """Normalize headers to simplify matching."""
        return (
            str(value).strip()
            .lower()
            .replace(" ", "_")
            .replace("(", "")
            .replace(")", "")
            .replace("/", "_")
            .replace("-", "_")
        )

    def post(self, request, *args, **kwargs):
        files = request.FILES.getlist("files")

        if not files:
            return JsonResponse({
                'status': 'error',
                'message': 'No files were uploaded.'
            })

        # Get or create session ID
        session_id = request.POST.get('session_id')
        if not session_id:
            session_id = str(uuid.uuid4())

        # Store files temporarily in cache (convert to byte strings)
        file_data_list = []
        for file in files:
            file_data_list.append({
                'name': file.name,
                'content': file.read(),
            })
        
        cache.set(f'upload_files_{session_id}', file_data_list, timeout=600)

        # Initialize progress
        cache.set(f'upload_progress_{session_id}', {
            'status': 'queued',
            'current_row': 0,
            'total_rows': 0,
            'current_file': 0,
            'total_files': len(files),
            'percentage': 0,
            'message': 'Upload queued...',
            'timestamp': time.time()
        }, timeout=600)

        # Return session ID immediately
        return JsonResponse({
            'status': 'queued',
            'session_id': session_id
        })

class ProcessUploadView(View):
    """
    Separate view that processes the upload.
    This is called via AJAX and streams progress updates.
    Merges data into existing reports if report_no and report_collection_date match.
    """
    
    @staticmethod
    def normalize_header(value):
        """Normalize headers to simplify matching."""
        return (
            str(value).strip()
            .lower()
            .replace(" ", "_")
            .replace("(", "")
            .replace(")", "")
            .replace("/", "_")
            .replace("-", "_")
        )
    
    @staticmethod
    def extract_report_metadata(ws, header_row=None, last_data_row=None):
        """Extract report metadata from the worksheet."""
        metadata = {
            'dti_office': None,
            'report_no': None,
            'report_collection_date': None,
            'certification': None,
            'name_of_collection_officer': None,
            'responsibility_center_code': None,
            'official_designation': None,
            'total': None,
            'undeposited_last_report': None,
            'collections_this_report': None,
        }
        
        print("\n" + "="*80)
        print("METADATA EXTRACTION DEBUG")
        print("="*80)
        print(f"Header row: {header_row}")
        print(f"Last data row (passed in): {last_data_row}")
        print(f"Worksheet max row: {ws.max_row}")
        print(f"Worksheet max column: {ws.max_column}")
        
        # Determine search boundary - only search BEFORE the data table header
        search_limit = header_row if header_row else min(10, ws.max_row + 1)
        
        # Search only header area for metadata (before data table)
        print(f"\n--- Searching HEADER area (rows 1-{search_limit-1}) ---")
        for row_idx in range(1, search_limit):
            for cell in ws[row_idx]:
                if not cell.value:
                    continue
                    
                cell_text = str(cell.value).strip()
                cell_lower = cell_text.lower()
                
                # Skip if this looks like the data table header row
                if 'official receipt' in cell_lower and 'number' in cell_lower:
                    continue
                
                # Extract DTI Office
                if 'dti' in cell_lower and 'office' in cell_lower:
                    metadata['dti_office'] = cell_text
                    print(f"✓ Found DTI Office at row {row_idx}: '{cell_text}'")
                
                # Extract Report No (but NOT Official Receipt No)
                if ('report no' in cell_lower or 'report_no' in cell_lower) and 'receipt' not in cell_lower and 'official' not in cell_lower:
                    # Check if number is in same cell
                    parts = cell_text.split('.')
                    if len(parts) > 1:
                        metadata['report_no'] = parts[-1].strip()
                        print(f"✓ Found Report No at row {row_idx}: '{metadata['report_no']}'")
                    else:
                        # Check adjacent cells
                        next_cell = ws.cell(row=cell.row, column=cell.column + 1)
                        if next_cell.value:
                            metadata['report_no'] = str(next_cell.value).strip()
                            print(f"✓ Found Report No at row {row_idx} (adjacent cell): '{metadata['report_no']}'")
                
                if 'responsibility center' in cell_lower or 'rc code' in cell_lower or 'rc_code' in cell_lower:
                # Check if value is in same cell (e.g., "RC Code: 1234")
                    parts = cell_text.split(':')
                    if len(parts) > 1:
                        metadata['responsibility_center_code'] = parts[-1].strip()
                        print(f"✓ Found RC Code at row {row_idx}: '{metadata['responsibility_center_code']}'")
                    else:
                        # Check adjacent cells (value might be in next cell)
                        next_cell = ws.cell(row=cell.row, column=cell.column + 1)
                        if next_cell.value:
                            metadata['responsibility_center_code'] = str(next_cell.value).strip()
                            print(f"✓ Found RC Code at row {row_idx} (adjacent cell): '{metadata['responsibility_center_code']}'")
                            
                # Extract Date (usually below DTI Office)
                if not metadata['report_collection_date']:
                    # Check if this looks like a date
                    if isinstance(cell.value, datetime.datetime):
                        metadata['report_collection_date'] = cell.value.date()
                        print(f"✓ Found Date at row {row_idx} (datetime): '{metadata['report_collection_date']}'")
                    elif isinstance(cell.value, str):
                        # Try to parse date formats
                        for fmt in ['%m/%d/%Y', '%m-%d-%Y', '%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%d-%b-%y', '%d-%b-%Y']:
                            try:
                                parsed_date = datetime.datetime.strptime(cell_text, fmt).date()
                                metadata['report_collection_date'] = parsed_date
                                print(f"✓ Found Date at row {row_idx} (string): '{metadata['report_collection_date']}'")
                                break
                            except:
                                continue
                    elif isinstance(cell.value, float) or isinstance(cell.value, int):
                        # Could be Excel date serial number
                        try:
                            if 30000 < cell.value < 60000:  # Valid Excel date range
                                base_date = datetime.datetime(1899, 12, 30)
                                parsed_date = base_date + datetime.timedelta(days=cell.value)
                                metadata['report_collection_date'] = parsed_date.date()
                                print(f"✓ Found Date at row {row_idx} (serial): '{metadata['report_collection_date']}'")
                        except:
                            pass
        
        # DYNAMIC FOOTER DETECTION: Find where the data table actually ends
        # Look for "CERTIFICATION" keyword or 3+ consecutive empty rows after header
        print(f"\n--- Detecting actual footer start ---")
        footer_start = None
        
        if header_row:
            # Strategy 1: Look for "CERTIFICATION" keyword (most reliable)
            for row_idx in range(header_row + 2, ws.max_row + 1):
                # Check all cells in the row (for merged cells spanning columns)
                row_text = []
                for cell in ws[row_idx]:
                    if cell.value:
                        row_text.append(str(cell.value).strip().lower())
                
                combined_text = ' '.join(row_text)
                if 'certification' in combined_text or 'i hereby certify' in combined_text:
                    footer_start = row_idx
                    print(f"✓ Found 'CERTIFICATION' marker at row {row_idx}")
                    break
            
            # Strategy 2: Look for consistent empty rows (fallback)
            if not footer_start:
                empty_count = 0
                for row_idx in range(header_row + 2, ws.max_row + 1):
                    # Check if row has any substantial data in the first few columns
                    has_data = False
                    for col_idx in range(1, min(6, ws.max_column + 1)):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        if cell.value:
                            cell_text = str(cell.value).strip()
                            # Ignore pure whitespace or underscores
                            if cell_text and cell_text not in ['_', '__', '___', '____', '_____', '______'] and not all(c in '_ ' for c in cell_text):
                                has_data = True
                                break
                    
                    if not has_data:
                        empty_count += 1
                        if empty_count >= 3:
                            footer_start = row_idx - 2  # Go back before the empty rows
                            print(f"✓ Detected footer start at row {footer_start} (3+ empty rows)")
                            break
                    else:
                        empty_count = 0
        
        # Fallback: Use a conservative estimate
        if not footer_start:
            footer_start = max(1, ws.max_row - 40)
            print(f"⚠ Using fallback footer start: row {footer_start}")
        
        print(f"Final footer_start: {footer_start}")
        print(f"\n--- Searching FOOTER area (rows {footer_start}-{ws.max_row}) ---")
        
        # Search for certification text (in footer area)
        certification_start = None
        
        for row_idx in range(footer_start, ws.max_row + 1):
            # Check ALL cells in row (for merged cells)
            row_text = []
            for cell in ws[row_idx]:
                if cell.value:
                    row_text.append(str(cell.value).strip().lower())
            
            combined_text = ' '.join(row_text)
            if 'i hereby certify' in combined_text or 'certification' in combined_text:
                certification_start = row_idx
                print(f"✓ Found Certification start at row {row_idx}")
                break
        
        if certification_start:
            # Collect certification text (next 10 rows)
            cert_lines = []
            for row_idx in range(certification_start, min(certification_start + 10, ws.max_row + 1)):
                row_text = []
                for cell in ws[row_idx]:
                    if cell.value:
                        row_text.append(str(cell.value).strip())
                if row_text:
                    cert_lines.append(' '.join(row_text))
            
            if cert_lines:
                metadata['certification'] = ' '.join(cert_lines)
                print(f"✓ Certification text: '{metadata['certification'][:100]}...'")
        
        # Search for collecting officer name and designation (in footer area)
        # Build a map of ALL non-empty cells in the footer for easier searching
        print(f"\n--- Building footer cell map ---")
        footer_cells = {}
        for row_idx in range(footer_start, ws.max_row + 1):
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value:
                    cell_text = str(cell.value).strip()
                    if cell_text and cell_text not in ['_', '__', '___', '____', '_____', '______']:
                        footer_cells[(row_idx, col_idx)] = cell_text
                        print(f"  Cell ({row_idx}, {col_idx}): '{cell_text}'")
        
        print(f"\n--- Searching for Name and Designation ---")
        # Now search for the labels
        for row_idx in range(footer_start, ws.max_row + 1):
            # Collect all values in this row (to handle merged cells)
            row_values = []
            for col_idx in range(1, ws.max_column + 1):
                if (row_idx, col_idx) in footer_cells:
                    row_values.append((col_idx, footer_cells[(row_idx, col_idx)]))
            
            # Check each cell in the row
            for col_idx, cell_text in row_values:
                cell_lower = cell_text.lower()
                
                # Look for "Name and Signature of Collecting Officer" label
                if 'name and signature of collecting officer' in cell_lower or 'name and signature' in cell_lower:
                    print(f"\n→ Found 'Name and Signature...' label at ({row_idx}, {col_idx})")
                    
                    # Strategy 1: Check the row ABOVE in the SAME column
                    if (row_idx - 1, col_idx) in footer_cells:
                        name_text = footer_cells[(row_idx - 1, col_idx)]
                        print(f"  Strategy 1: Checking ({row_idx - 1}, {col_idx}): '{name_text}'")
                        if 'name and signature' not in name_text.lower():
                            metadata['name_of_collection_officer'] = name_text
                            print(f"  ✓ FOUND NAME: '{name_text}'")
                    
                    # Strategy 2: Check 2 rows above
                    if not metadata['name_of_collection_officer'] and (row_idx - 2, col_idx) in footer_cells:
                        name_text = footer_cells[(row_idx - 2, col_idx)]
                        print(f"  Strategy 2: Checking ({row_idx - 2}, {col_idx}): '{name_text}'")
                        if 'name and signature' not in name_text.lower():
                            metadata['name_of_collection_officer'] = name_text
                            print(f"  ✓ FOUND NAME: '{name_text}'")
                    
                    # Strategy 3: Search ALL cells above in ALL columns (for merged cells)
                    if not metadata['name_of_collection_officer']:
                        print(f"  Strategy 3: Searching all columns in rows above...")
                        for check_row in range(max(footer_start, row_idx - 3), row_idx):
                            for check_col in range(1, ws.max_column + 1):
                                if (check_row, check_col) in footer_cells:
                                    potential_name = footer_cells[(check_row, check_col)]
                                    potential_lower = potential_name.lower()
                                    # Check if this looks like a name (not a label)
                                    if (potential_name and 
                                        'name and signature' not in potential_lower and
                                        'official designation' not in potential_lower and
                                        'certification' not in potential_lower and
                                        'certify' not in potential_lower and
                                        'summary' not in potential_lower and
                                        'total' not in potential_lower and
                                        len(potential_name) > 5):
                                        print(f"    Checking ({check_row}, {check_col}): '{potential_name}' - MATCH!")
                                        metadata['name_of_collection_officer'] = potential_name
                                        break
                            if metadata['name_of_collection_officer']:
                                break
                    
                    if not metadata['name_of_collection_officer']:
                        print(f"  ✗ Could not find name")
                
                # Look for "Official Designation" label
                if 'official designation' in cell_lower:
                    print(f"\n→ Found 'Official Designation' label at ({row_idx}, {col_idx})")
                    
                    # Strategy 1: Check the row ABOVE in the SAME column
                    if (row_idx - 1, col_idx) in footer_cells:
                        designation_text = footer_cells[(row_idx - 1, col_idx)]
                        print(f"  Strategy 1: Checking ({row_idx - 1}, {col_idx}): '{designation_text}'")
                        if 'official designation' not in designation_text.lower():
                            metadata['official_designation'] = designation_text
                            print(f"  ✓ FOUND DESIGNATION: '{designation_text}'")
                    
                    # Strategy 2: Check 2 rows above
                    if not metadata['official_designation'] and (row_idx - 2, col_idx) in footer_cells:
                        designation_text = footer_cells[(row_idx - 2, col_idx)]
                        print(f"  Strategy 2: Checking ({row_idx - 2}, {col_idx}): '{designation_text}'")
                        if 'official designation' not in designation_text.lower():
                            metadata['official_designation'] = designation_text
                            print(f"  ✓ FOUND DESIGNATION: '{designation_text}'")
                    
                    # Strategy 3: Check SAME ROW in ALL columns to the LEFT (merged cell scenario)
                    if not metadata['official_designation']:
                        print(f"  Strategy 3: Checking same row, all columns to the left...")
                        for check_col in range(1, col_idx):
                            if (row_idx, check_col) in footer_cells:
                                designation_text = footer_cells[(row_idx, check_col)]
                                designation_lower = designation_text.lower()
                                print(f"    Checking ({row_idx}, {check_col}): '{designation_text}'")
                                # Make sure it's not the label itself
                                if ('official designation' not in designation_lower and
                                    'name and signature' not in designation_lower and
                                    'certification' not in designation_lower and
                                    len(designation_text) > 3):
                                    metadata['official_designation'] = designation_text
                                    print(f"    ✓ FOUND DESIGNATION: '{designation_text}'")
                                    break
                    
                    # Strategy 4: Check column to the RIGHT
                    if not metadata['official_designation'] and (row_idx, col_idx + 1) in footer_cells:
                        designation_text = footer_cells[(row_idx, col_idx + 1)]
                        print(f"  Strategy 4: Checking ({row_idx}, {col_idx + 1}): '{designation_text}'")
                        if 'official designation' not in designation_text.lower():
                            metadata['official_designation'] = designation_text
                            print(f"  ✓ FOUND DESIGNATION: '{designation_text}'")
                    
                    # Strategy 5: Search ALL nearby cells (within 3 rows above/below and all columns)
                    if not metadata['official_designation']:
                        print(f"  Strategy 5: Searching all nearby cells...")
                        for check_row in range(max(footer_start, row_idx - 3), min(ws.max_row + 1, row_idx + 2)):
                            for check_col in range(1, ws.max_column + 1):
                                if (check_row, check_col) in footer_cells:
                                    potential_designation = footer_cells[(check_row, check_col)]
                                    potential_lower = potential_designation.lower()
                                    # Check if this looks like a designation
                                    if (potential_designation and 
                                        'official designation' not in potential_lower and
                                        'name and signature' not in potential_lower and
                                        'certification' not in potential_lower and
                                        'certify' not in potential_lower and
                                        'summary' not in potential_lower and
                                        'total' not in potential_lower and
                                        len(potential_designation) > 5):
                                        print(f"    Checking ({check_row}, {check_col}): '{potential_designation}' - MATCH!")
                                        metadata['official_designation'] = potential_designation
                                        break
                            if metadata['official_designation']:
                                break
                    
                    if not metadata['official_designation']:
                        print(f"  ✗ Could not find designation")
                
                # Look for "Special Collecting Officer" as a standalone value (fallback)
                if cell_lower == 'special collecting officer' and not metadata['official_designation']:
                    metadata['official_designation'] = 'Special Collecting Officer'
                    print(f"✓ Found standalone 'Special Collecting Officer' at ({row_idx}, {col_idx})")
        
        print("\n" + "="*80)
        print("FINAL METADATA RESULTS:")
        print("="*80)
        for key, value in metadata.items():
            print(f"{key}: {value}")
        print("="*80 + "\n")
        
        return metadata
    
    def get(self, request, session_id):   
        def process_and_stream():
            # Track created reports/items for rollback on cancel
            created_report_ids = []
            created_item_ids = []
            merged_report_ids = []
            
            try:
                # Check for cancellation at the start
                if cache.get(f'upload_cancel_{session_id}'):
                    yield self._sse_message({
                        'status': 'cancelled',
                        'message': 'Upload cancelled by user'
                    })
                    cache.delete(f'upload_cancel_{session_id}')
                    return
                
                # Get files from cache
                file_data_list = cache.get(f'upload_files_{session_id}')
                if not file_data_list:
                    yield self._sse_message({'status': 'error', 'message': 'Files not found'})
                    return

                yield self._sse_message({
                    'status': 'processing',
                    'message': 'Starting processing...',
                    'percentage': 0
                })

                created_reports = []
                merged_reports_list = []
                
                # First pass: Count total rows
                total_rows = 0
                file_row_counts = []
                
                yield self._sse_message({
                    'status': 'counting',
                    'message': 'Counting rows...',
                    'percentage': 0
                })
                
                for file_data in file_data_list:
                    # Check for cancellation during counting
                    if cache.get(f'upload_cancel_{session_id}'):
                        yield self._sse_message({
                            'status': 'cancelled',
                            'message': 'Upload cancelled - no data was created'
                        })
                        cache.delete(f'upload_cancel_{session_id}')
                        return
                    
                    ext = os.path.splitext(file_data['name'])[1].lower()
                    if file_data['name'].startswith("~$") or ext not in [".xlsx", ".xls"]:
                        file_row_counts.append(0)
                        continue
                    
                    try:
                        file_obj = BytesIO(file_data['content'])
                        wb = load_workbook(file_obj, data_only=True)
                        ws = wb.active
                        row_count = max(0, ws.max_row - 2)
                        file_row_counts.append(row_count)
                        total_rows += row_count
                        wb.close()
                    except Exception as e:
                        print(f"Error counting rows: {e}")
                        file_row_counts.append(0)

                yield self._sse_message({
                    'status': 'processing',
                    'current_row': 0,
                    'total_rows': total_rows,
                    'current_file': 0,
                    'total_files': len(file_data_list),
                    'percentage': 0,
                    'message': f'Processing {total_rows} rows...'
                })

                current_row = 0
                
                # Process each file
                for file_index, file_data in enumerate(file_data_list):
                    # Check for cancellation before each file
                    if cache.get(f'upload_cancel_{session_id}'):
                        # Rollback all created data
                        self._rollback_changes(created_report_ids, created_item_ids, merged_report_ids)
                        yield self._sse_message({
                            'status': 'cancelled',
                            'message': 'Upload cancelled - all data has been deleted'
                        })
                        cache.delete(f'upload_cancel_{session_id}')
                        return
                    
                    ext = os.path.splitext(file_data['name'])[1].lower()
                    
                    yield self._sse_message({
                        'status': 'processing',
                        'current_row': current_row,
                        'total_rows': total_rows,
                        'current_file': file_index + 1,
                        'total_files': len(file_data_list),
                        'percentage': round((current_row / total_rows * 100) if total_rows > 0 else 0, 2),
                        'message': f'Processing {file_data["name"]}...',
                        'current_filename': file_data['name']
                    })
                    
                    if file_data['name'].startswith("~$") or ext not in [".xlsx", ".xls"]:
                        continue

                    try:
                        file_obj = BytesIO(file_data['content'])
                        wb = load_workbook(file_obj, data_only=True)
                        ws = wb.active
                    except Exception as e:
                        print(f"Error opening file: {e}")
                        continue

                    # Auto-detect header row
                    header_row = None
                    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
                        cells = [str(c).strip().lower() for c in row if c]
                        if any("official receipt" in c or "date" == c for c in cells):
                            header_row = i
                            break

                    if not header_row:
                        wb.close()
                        continue

                    # Process all data rows first to find where they end
                    last_data_row = header_row + 2
                    for row_index, row in enumerate(ws.iter_rows(min_row=header_row + 2, values_only=True), start=header_row + 2):
                        if any(row):
                            last_data_row = row_index
                    
                    # Extract metadata
                    metadata = self.extract_report_metadata(ws, header_row, last_data_row)

                    # Combine header rows
                    headers = []
                    for c1, c2 in zip(ws[header_row], ws[header_row + 1]):
                        part1 = str(c1.value).strip() if c1.value else ""
                        part2 = str(c2.value).strip() if c2.value else ""
                        combined = f"{part1} {part2}".strip()
                        headers.append(self.normalize_header(combined))

                    # Field mapping
                    field_map = {}
                    for f in CollectionReportItem._meta.fields:
                        opts = {self.normalize_header(f.verbose_name or ""), self.normalize_header(f.name)}
                        for h in headers:
                            if h in opts:
                                field_map[h] = f.name

                    # Manual mappings
                    if "official_receipt_number" in headers:
                        field_map["official_receipt_number"] = "number"
                    if "official_receipt_number" not in field_map and "number" in headers:
                        field_map["number"] = "number"
                    if "number" not in field_map:
                        for h in headers:
                            if "receipt" in h and "number" in h:
                                field_map[h] = "number"
                                break
                    if "number" not in field_map:
                        for idx, h in enumerate(headers):
                            if h == "number":
                                field_map[h] = "number"
                                break
                    
                    if "date" not in field_map and any("date" in h for h in headers):
                        date_header = next(h for h in headers if "date" in h)
                        field_map[date_header] = "date"
                    
                    # Map responsibility center code field - ENHANCED VERSION
                    print(f"\nDEBUG - All headers after normalization: {headers}")
                    print(f"DEBUG - Looking for RC code mapping...")

                    rc_mapped = False
                    for idx, h in enumerate(headers):
                        if not rc_mapped:
                            print(f"  Checking header[{idx}]: '{h}'")
                            
                            # Strategy 1: Direct match
                            if h in ["rc_code", "responsibility_center_code"]:
                                field_map[h] = "rc_code"
                                print(f"  ✓ Mapped '{h}' to 'rc_code' (direct match)")
                                rc_mapped = True
                                break
                            
                            # Strategy 2: Contains both "responsibility" and "center"
                            if "responsibility" in h and "center" in h:
                                field_map[h] = "rc_code"
                                print(f"  ✓ Mapped '{h}' to 'rc_code' (contains 'responsibility' + 'center')")
                                rc_mapped = True
                                break
                            
                            # Strategy 3: Contains "responsibility" and "code"
                            if "responsibility" in h and "code" in h:
                                field_map[h] = "rc_code"
                                print(f"  ✓ Mapped '{h}' to 'rc_code' (contains 'responsibility' + 'code')")
                                rc_mapped = True
                                break

                    if not rc_mapped:
                        print("  ✗ No RC code column found in headers")
                    else:
                        print(f"DEBUG - Field map now includes RC code mapping")

                    print(f"DEBUG - Complete field_map: {field_map}\n")

                    # Check for existing report and merge
                    is_merge = False
                    report = None
                    
                    if metadata['report_no']:
                        try:
                            report = CollectionReport.objects.filter(
                                report_no=metadata['report_no']
                            ).first()
                            
                            if report:
                                is_merge = True
                                if report.id not in merged_report_ids:
                                    merged_report_ids.append(report.id)
                                    # Store in cache for cleanup
                                    cache.set(f'upload_merged_reports_{session_id}', merged_report_ids, timeout=600)
                                print(f"DEBUG - Found existing report by report_no: {report.report_no}")
                                
                                yield self._sse_message({
                                    'status': 'processing',
                                    'message': f'Merging into existing report {report.report_no}...',
                                    'current_filename': file_data['name']
                                })
                                
                                # Update metadata if provided
                                if metadata['report_collection_date'] and not report.report_collection_date:
                                    report.report_collection_date = metadata['report_collection_date']
                                if metadata['dti_office'] and not report.dti_office:
                                    report.dti_office = metadata['dti_office']
                                if metadata['certification'] and not report.certification:
                                    report.certification = metadata['certification']
                                if metadata['name_of_collection_officer'] and not report.name_of_collection_officer:
                                    report.name_of_collection_officer = metadata['name_of_collection_officer']
                                if metadata['official_designation'] and not report.official_designation:
                                    report.official_designation = metadata['official_designation']
                                report.save()
                            
                        except Exception as e:
                            print(f"Error checking for existing report: {e}")
                    
                    # Fallback check by date
                    if not report and metadata['report_collection_date']:
                        try:
                            report = CollectionReport.objects.filter(
                                report_collection_date=metadata['report_collection_date']
                            ).first()
                            
                            if report:
                                is_merge = True
                                if report.id not in merged_report_ids:
                                    merged_report_ids.append(report.id)
                                    cache.set(f'upload_merged_reports_{session_id}', merged_report_ids, timeout=600)
                                print(f"DEBUG - Found existing report by date: {report.report_collection_date}")
                                
                                yield self._sse_message({
                                    'status': 'processing',
                                    'message': f'Merging into existing report for {report.report_collection_date}...',
                                    'current_filename': file_data['name']
                                })
                        except Exception as e:
                            print(f"Error checking for existing report by date: {e}")
                    
                    # Create new report only if no existing one found
                    if not report:
                        report = CollectionReport.objects.create(
                            dti_office=metadata['dti_office'],
                            report_no=metadata['report_no'],
                            report_collection_date=metadata['report_collection_date'],
                            certification=metadata['certification'],
                            name_of_collection_officer=metadata['name_of_collection_officer'],
                            official_designation=metadata['official_designation'],
                            responsibility_center_code=metadata.get('responsibility_center_code', None)
                        )
                        created_report_ids.append(report.id)
                        # Store in cache for cleanup
                        cache.set(f'upload_created_reports_{session_id}', created_report_ids, timeout=600)
                        print(f"DEBUG - Created new report: {report.report_no} ({report.report_collection_date})")
                    
                    db_column_indices = [idx for idx, h in enumerate(headers) if h in field_map]
                    empty_row_count = 0
                    MAX_EMPTY_ROWS = 5
                    items_added = 0
                    row_check_counter = 0

                    # Process rows
                    for row_index, row in enumerate(ws.iter_rows(min_row=header_row + 2, values_only=True), start=header_row + 2):
                        # Check for cancellation every 10 rows
                        row_check_counter += 1
                        if row_check_counter % 10 == 0:
                            if cache.get(f'upload_cancel_{session_id}'):
                                wb.close()
                                self._rollback_changes(created_report_ids, created_item_ids, merged_report_ids)
                                yield self._sse_message({
                                    'status': 'cancelled',
                                    'message': 'Upload cancelled - all data has been deleted'
                                })
                                cache.delete(f'upload_cancel_{session_id}')
                                return
                        
                        if not any(row[idx] if idx < len(row) else None for idx in db_column_indices):
                            empty_row_count += 1
                            if empty_row_count >= MAX_EMPTY_ROWS:
                                break
                            continue

                        empty_row_count = 0
                        current_row += 1
                        
                        # Send update every 5 rows
                        if current_row % 5 == 0:
                            percentage = round((current_row / total_rows * 100) if total_rows > 0 else 0, 2)
                            yield self._sse_message({
                                'status': 'processing',
                                'current_row': current_row,
                                'total_rows': total_rows,
                                'current_file': file_index + 1,
                                'total_files': len(file_data_list),
                                'percentage': percentage,
                                'message': f'Processing row {current_row} of {total_rows}',
                                'current_filename': file_data['name']
                            })

                        data = {}
                        for idx, value in enumerate(row):
                            header = headers[idx] if idx < len(headers) else None
                            field_name = field_map.get(header)
                            if not field_name:
                                continue

                            # Handle Date column
                            if field_name == "date":
                                if isinstance(value, datetime.datetime):
                                    value = value.date()
                                elif isinstance(value, float):
                                    base_date = datetime.datetime(1899, 12, 30)
                                    value = base_date + datetime.timedelta(days=value)
                                    value = value.date()
                                elif isinstance(value, str):
                                    try:
                                        value = datetime.datetime.strptime(value.strip(), "%Y-%m-%d").date()
                                    except:
                                        try:
                                            value = datetime.datetime.strptime(value.strip(), "%m/%d/%Y").date()
                                        except:
                                            value = None

                            data[field_name] = value
                        
                            if field_name == "rc_code" and value:
                                print(f"DEBUG - Found RC code in row {row_index}: '{value}'")
                                
                        # Receipt number fallback logic
                        if 'number' not in data or not data['number']:
                            date_col_idx = None
                            for idx, h in enumerate(headers):
                                if field_map.get(h) == 'date':
                                    date_col_idx = idx
                                    break
                            
                            for idx, value in enumerate(row):
                                if not value:
                                    continue
                                
                                if date_col_idx is not None and idx == date_col_idx:
                                    continue
                                    
                                value_str = str(value).strip()
                                
                                if not value_str or value_str.lower() in ['none', 'n/a', '']:
                                    continue
                                
                                if value_str.startswith(('BN-', 'RN-', 'OR-', 'AR-')):
                                    data['number'] = value_str
                                    break
                                
                                if value_str.count('-') >= 2 and len(value_str) > 10:
                                    data['number'] = value_str
                                    break
                            
                            if 'number' not in data or not data['number']:
                                if date_col_idx is not None:
                                    for offset in [1, 2]:
                                        check_idx = date_col_idx + offset
                                        if check_idx < len(row):
                                            value = row[check_idx]
                                            if not value:
                                                continue
                                            
                                            value_str = str(value).strip()
                                            
                                            try:
                                                if isinstance(value, (int, float)):
                                                    if value == int(value) and int(value) > 100000:
                                                        data['number'] = str(int(value))
                                                        break
                                                elif value_str.isdigit() and len(value_str) >= 6:
                                                    data['number'] = value_str
                                                    break
                                            except:
                                                continue

                        if not data:
                            continue

                        # Prevent duplicate items when merging
                        if is_merge and 'number' in data and data['number']:
                            existing_item = report.report_items.filter(
                                number=data['number'],
                                date=data.get('date')
                            ).first()
                            
                            if existing_item:
                                print(f"DEBUG - Skipping duplicate item: {data['number']} on {data.get('date')}")
                                continue

                        try:
                            with transaction.atomic():
                                item = CollectionReportItem.objects.create(**data)
                                created_item_ids.append(item.id)
                                cache.set(f'upload_created_items_{session_id}', created_item_ids, timeout=600)
                                report.report_items.add(item)
                                items_added += 1
                                
                        except Exception as e:
                            print(f"Failed to save row {row_index}: {e}")
                            continue
                    
                    # After all items are added, populate responsibility_center_code
                    if not report.responsibility_center_code:
                        first_item = report.report_items.first()
                        if first_item and first_item.rc_code:
                            report.responsibility_center_code = first_item.rc_code
                            report.save()
                            print(f"DEBUG: Auto-set responsibility_center_code to '{report.responsibility_center_code}'")
                        else:
                            print(f"DEBUG: Could not auto-set RC code - first item: {first_item}, rc_code: {first_item.rc_code if first_item else 'N/A'}")
                        
                    # Track reports
                    if is_merge:
                        if report not in merged_reports_list:
                            merged_reports_list.append(report)
                        print(f"DEBUG - Merged {items_added} items into report {report.report_no}")
                    else:
                        created_reports.append(report)
                        print(f"DEBUG - Created report {report.report_no} with {items_added} items")
                    
                    wb.close()

                # Determine redirect URL
                all_reports = created_reports + merged_reports_list
                if len(all_reports) > 1:
                    redirect_url = reverse("collection-report-list")
                elif all_reports:
                    redirect_url = reverse("collection-report", args=[all_reports[-1].pk])
                else:
                    redirect_url = reverse("all-documents")
                
                # Send 100% progress
                yield self._sse_message({
                    'status': 'processing',
                    'current_row': total_rows,
                    'total_rows': total_rows,
                    'current_file': len(file_data_list),
                    'total_files': len(file_data_list),
                    'percentage': 100,
                    'message': 'Finalizing...'
                })
                
                time.sleep(0.3)
                
                # Build completion message
                completion_parts = []
                if created_reports:
                    completion_parts.append(f'{len(created_reports)} new report(s) created')
                if merged_reports_list:
                    completion_parts.append(f'{len(merged_reports_list)} report(s) updated')
                
                completion_message = ' and '.join(completion_parts) if completion_parts else 'Upload complete!'
                
                # Send completion
                yield self._sse_message({
                    'status': 'complete',
                    'current_row': total_rows,
                    'total_rows': total_rows,
                    'current_file': len(file_data_list),
                    'total_files': len(file_data_list),
                    'percentage': 100,
                    'message': completion_message,
                    'redirect_url': redirect_url
                })

                # Clean up cache
                cache.delete(f'upload_files_{session_id}')
                cache.delete(f'upload_created_reports_{session_id}')
                cache.delete(f'upload_merged_reports_{session_id}')
                cache.delete(f'upload_created_items_{session_id}')

            except Exception as e:
                print(f"Processing error: {e}")
                import traceback
                traceback.print_exc()
                yield self._sse_message({
                    'status': 'error',
                    'message': f'Error: {str(e)}'
                })
        
        response = StreamingHttpResponse(
            process_and_stream(),
            content_type='text/event-stream'
        )
        response['Cache-Control'] = 'no-cache'
        response['X-Accel-Buffering'] = 'no'
        return response
    
    def _rollback_changes(self, created_report_ids, created_item_ids, merged_report_ids):
        """Rollback all changes made during this upload session"""
        try:
            # Delete newly created reports
            if created_report_ids:
                from ..models.collection_models import CollectionReport
                deleted_count = CollectionReport.objects.filter(id__in=created_report_ids).delete()
                print(f"Rolled back {deleted_count[0]} created reports")
            
            # Delete items added to merged reports
            if created_item_ids:
                from ..models.collection_models import CollectionReportItem
                deleted_count = CollectionReportItem.objects.filter(id__in=created_item_ids).delete()
                print(f"Rolled back {deleted_count[0]} created items")
        except Exception as e:
            print(f"Error during rollback: {e}")
    
    def _sse_message(self, data):
        """Format data as SSE message"""
        return f"data: {json.dumps(data)}\n\n"


class CancelUploadView(View):
    """
    Cancel an ongoing upload and clean up all created data
    """
    
    def post(self, request, session_id):
        try:
            # Set cancellation flag immediately
            cache.set(f'upload_cancel_{session_id}', True, timeout=600)
            
            # Get the list of reports created/modified during this session
            created_report_ids = cache.get(f'upload_created_reports_{session_id}', [])
            merged_report_ids = cache.get(f'upload_merged_reports_{session_id}', [])
            created_item_ids = cache.get(f'upload_created_items_{session_id}', [])
            
            deleted_counts = {
                'reports': 0,
                'items': 0,
                'merged_reports': 0
            }
            
            # Delete all newly created reports
            if created_report_ids:
                from ..models.collection_models import CollectionReport
                deleted_reports = CollectionReport.objects.filter(
                    id__in=created_report_ids
                ).delete()
                deleted_counts['reports'] = deleted_reports[0] if deleted_reports else 0
                print(f"Deleted {deleted_counts['reports']} newly created reports")
            
            # For merged reports, only delete the items that were added in this session
            if created_item_ids:
                from ..models.collection_models import CollectionReportItem
                deleted_items = CollectionReportItem.objects.filter(
                    id__in=created_item_ids
                ).delete()
                deleted_counts['items'] = deleted_items[0] if deleted_items else 0
                print(f"Deleted {deleted_counts['items']} items from merged reports")
                deleted_counts['merged_reports'] = len(merged_report_ids)
            
            # Clean up cache
            cache.delete(f'upload_files_{session_id}')
            cache.delete(f'upload_progress_{session_id}')
            cache.delete(f'upload_created_reports_{session_id}')
            cache.delete(f'upload_merged_reports_{session_id}')
            cache.delete(f'upload_created_items_{session_id}')
            
            return JsonResponse({
                'status': 'cancelled',
                'message': 'Upload cancelled and all data deleted',
                'deleted_counts': deleted_counts
            })
            
        except Exception as e:
            print(f"Error during cancellation cleanup: {e}")
            import traceback
            traceback.print_exc()
            
            # Still set the cancel flag even if cleanup fails
            cache.set(f'upload_cancel_{session_id}', True, timeout=600)
            
            return JsonResponse({
                'status': 'error',
                'message': f'Error during cancellation: {str(e)}'
            })
    
class UploadProgressStreamView(View):
    """Legacy view - kept for backwards compatibility but not used in new flow"""
    
    def get(self, request, session_id):
        progress_data = cache.get(f'upload_progress_{session_id}')
        return JsonResponse(progress_data or {'status': 'not_found'})
